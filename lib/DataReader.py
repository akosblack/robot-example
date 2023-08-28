from openpyxl import load_workbook
import os
import requests
import csv
import re
import logging
from pathlib import Path


class DataReader:
    def __init__(self):
        # Log fájl elérési útja és neve
        #Path.cwd()
        log_folder = Path.cwd() / 'log'
        self.log_file_path = log_folder / 'log_file.log'

        self.data_folder = os.path.join(os.path.dirname(__file__), "..", "input")
        
        # Log fájl beállítása
        logging.basicConfig(level=logging.DEBUG, filename=self.log_file_path, filemode='w',
                            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        
        self.logger = logging.getLogger(__name__)  # Itt hozzuk létre az osztályváltozót
        self.logger.info('Logger is set up')

    def read_excel(self, file_path):
        data = []
        workbook = load_workbook(file_path)
        sheet = workbook.active

        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header_cell, value in zip(header, row):
                if header_cell is not None and value is not None:
                    row_data[header_cell] = value
            if row_data:
                data.append(row_data)

        workbook.close()
        return data
    
    def download_file(self, url, destination):
        response = requests.get(url)
        if response.status_code == 200:
            os.makedirs(os.path.dirname(destination), exist_ok=True)
            with open(destination, "wb") as f:
                f.write(response.content)
            return True
        return False

    def append_to_csv(self, path, content):
        with open(path, 'a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(content)

    def append_to_txt(self, file_path, content):
        if not os.path.exists(file_path):
            with open(file_path, 'w') as file:
                file.write(content + '\n')
        with open(file_path, 'a') as file:
            file.write(content + '\n')


    def extract_invoice_info(self, text, id, due_date):
        lines = text.strip().split('\n')
        info = {
            'InvoiceNo': None,
            'InvoiceDate': None,
            'CompanyName': None,
            'TotalDue': None
        }

        index = 0

        for line in lines:
            line = line.strip()

            if line.startswith('Total'):
                self.append_to_txt(self.log_file.path, f"{id} Total line: {line}")
                formatted_line = line.replace('Total', '')
                formatted_line = line.replace(':', '')
                formatted_line = line.replace(',', '')
                formatted_line = formatted_line.replace('Total', '')
                formatted_line = formatted_line.replace(' ', '')
                formatted_line = formatted_line.replace('$', '')
                info['TotalDue'] = formatted_line

            # Keresés a mintára: SZÁM.SZÁM
            pattern = r'\d+\.\d+'

            # Találatok kinyerése a szövegből
            matches = re.findall(pattern, line)

            print("Matches:", matches)
            self.append_to_txt('log_file_path', f"{id} Matches: {matches}")
            
            try:
                if line.startswith('#'):
                    if 'tem Description' in line:
                        donothing = 1
                        #break
                    else:
                        info['InvoiceNo'] = line.split('#', 1)[1].strip()
                        
                elif line.startswith('# '):
                    if 'tem Description' in line:
                        break
                    else:
                        info['InvoiceNo'] = line.split('# ', 1)[1].strip()
                
                elif line.startswith('Buckley'):
                    formatted_line = line.replace('Buckley, Washington Invoice #', '')
                    info['InvoiceNo'] = formatted_line.strip()
            except:
                logging.error(f"Error while extracting InvoiceNo from {line}")
              
            if line.startswith('Sit Amet Corp.'):
                info['CompanyName'] = 'Sit Amet Corp.'
            elif line.startswith('Aenean LLC'):
                info['CompanyName'] = 'Aenean LLC'
         

            index += 1  # Sor index növelése minden iterációban
        
        print(f"Extracted info: {info}")
        return f"{id},{due_date},{info['InvoiceNo']},{info['CompanyName']},{info['TotalDue']}"

    def rewrite_first_line(self,input_txt_file,output_csv_file):
        # Beolvassuk az eredeti .txt fájlt
        with open(input_txt_file, 'r') as file:
            lines = file.readlines()

        # Módosítjuk az első sort
        new_first_row = 'ID,DueDate,InvoiceNo,InvoiceDate,CompanyName,TotalDue\n'
        lines[0] = new_first_row
        lines_without_dashes = [line.replace(':', '') for line in lines]

        # Az módosított tartalmat .csv fájlként elmentjük
        with open(output_csv_file, "w") as file:
            file.writelines(lines_without_dashes)


